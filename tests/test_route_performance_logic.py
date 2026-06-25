import csv
import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MODULE_PATH = ROOT / "src" / "30_route_performance.py"
if not MODULE_PATH.exists():
    MODULE_PATH = ROOT / "work" / "30_route_performance.py"
spec = importlib.util.spec_from_file_location("route_performance30", MODULE_PATH)
route_performance = importlib.util.module_from_spec(spec)
sys.modules[spec.name] = route_performance
spec.loader.exec_module(route_performance)


def write_route(path: Path, count: int = 5, step: float = 0.001) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        for i in range(count):
            row = [""] * 16
            row[14] = 139.0 + i * step
            row[15] = 35.0
            writer.writerow(row)


def write_trip(path: Path, coords: list[tuple[str, float]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        for idx, (time_text, lon) in enumerate(coords):
            row = [""] * 16
            row[2] = "20250102"
            row[6] = time_text
            row[8] = "trip-1"
            row[14] = lon
            row[15] = 35.0
            writer.writerow(row)


class RoutePerformanceLogicTest(unittest.TestCase):
    def test_coarse_gps_segment_fills_intermediate_buckets(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp)
            route_dir = project / route_performance.ROUTE_DIR_CANDIDATES[0]
            trip_dir = project / route_performance.SCREENING_DIR_CANDIDATES[0]
            route_dir.mkdir(parents=True)
            trip_dir.mkdir(parents=True)
            write_route(route_dir / "route_a.csv", step=0.0005)
            write_trip(trip_dir / "trip.csv", [("08:00:00", 139.0), ("08:04:00", 139.002)])

            result = route_performance.analyze_project(project, allowed_dates={"20250102"}, allowed_hours={8})

            self.assertEqual(result["results"][0]["events"], 4)

    def test_same_trip_is_not_counted_twice_in_same_bucket(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp)
            route_dir = project / route_performance.ROUTE_DIR_CANDIDATES[0]
            trip_dir = project / route_performance.SCREENING_DIR_CANDIDATES[0]
            route_dir.mkdir(parents=True)
            trip_dir.mkdir(parents=True)
            write_route(route_dir / "route_a.csv", count=3)
            write_trip(
                trip_dir / "trip.csv",
                [
                    ("08:00:00", 139.0),
                    ("08:01:00", 139.001),
                    ("08:02:00", 139.0),
                    ("08:03:00", 139.001),
                ],
            )

            result = route_performance.analyze_project(project, allowed_dates={"20250102"}, allowed_hours={8})

            self.assertEqual(result["results"][0]["events_csv"], "")
            daily_csv = Path(result["results"][0]["daily_hourly_csv"])
            with daily_csv.open("r", encoding="utf-8-sig", newline="") as fh:
                rows = list(csv.DictReader(fh))
            bucket_one_rows = [row for row in rows if row["bucket_index"] == "1" and row["date"] == "20250102"]
            self.assertTrue(bucket_one_rows)
            self.assertTrue(all(row["trip_count"] == "1" for row in bucket_one_rows))

    def test_daily_hourly_summary_and_viewer_can_be_rebuilt_later(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp)
            route_dir = project / route_performance.ROUTE_DIR_CANDIDATES[0]
            trip_dir = project / route_performance.SCREENING_DIR_CANDIDATES[0]
            route_dir.mkdir(parents=True)
            trip_dir.mkdir(parents=True)
            write_route(route_dir / "route_a.csv", step=0.0005)
            write_trip(trip_dir / "trip.csv", [("08:00:00", 139.0), ("08:04:00", 139.002)])

            result = route_performance.analyze_project(project, allowed_dates={"20250102"}, allowed_hours={8})
            daily_csv = Path(result["results"][0]["daily_hourly_csv"])
            daily_xlsx = Path(result["results"][0]["daily_xlsx_files"][0])
            rebuilt_viewer = route_performance.build_viewer_from_output(result["output_dir"])

            self.assertTrue(daily_csv.exists())
            self.assertEqual(result["results"][0]["events_csv"], "")
            self.assertTrue(daily_xlsx.exists())
            with daily_csv.open("r", encoding="utf-8-sig", newline="") as fh:
                rows = list(csv.DictReader(fh))
            self.assertTrue(rows)
            self.assertTrue(all(row["date"] == "20250102" for row in rows))
            self.assertIn("freeflow_speed_kmh", rows[0])
            self.assertIn("congested_speed_kmh", rows[0])
            wb = load_workbook(daily_xlsx, read_only=True)
            self.assertIn("speed_forward", wb.sheetnames)
            self.assertIn("speed_free_forward", wb.sheetnames)
            self.assertIn("speed_jam_forward", wb.sheetnames)
            self.assertIn("trip_forward", wb.sheetnames)
            self.assertIn("volume_forward", wb.sheetnames)
            self.assertNotIn("speed3h_forward", wb.sheetnames)
            self.assertNotIn("volume3h_forward", wb.sheetnames)
            self.assertEqual(wb["speed_forward"].max_column, 29)
            self.assertEqual(wb["speed_free_forward"].max_column, 29)
            self.assertEqual(wb["speed_jam_forward"].max_column, 29)
            self.assertEqual(wb["trip_forward"].max_column, 29)
            self.assertEqual(wb["volume_forward"].max_column, 29)
            wb.close()
            viewer_html = rebuilt_viewer.read_text(encoding="utf-8")
            self.assertIn("state.hours", viewer_html)
            self.assertIn("redrawButton", viewer_html)
            self.assertIn("exportButton", viewer_html)
            self.assertIn("hoursList", viewer_html)
            self.assertIn("onkeydown", viewer_html)
            self.assertIn("ArrowUp", viewer_html)
            self.assertIn("ArrowDown", viewer_html)
            self.assertIn("HAS_LEAFLET", viewer_html)
            self.assertIn("initFallbackMap", viewer_html)
            self.assertIn("背景地図なし / ルート形状のみ", viewer_html)
            self.assertIn("speedKind", viewer_html)
            self.assertIn("exportWorkbook", viewer_html)
            self.assertIn("抽出条件", viewer_html)
            self.assertIn("SPEED_BREAKS", viewer_html)
            self.assertIn("TRIP_COLORS", viewer_html)
            self.assertIn("VOLUME_COLORS", viewer_html)
            self.assertIn("トリップ", viewer_html)
            self.assertIn("交通量", viewer_html)
            self.assertIn("平均所要時間(分)", viewer_html)
            self.assertIn("routeTravelTime", viewer_html)
            self.assertIn("segmentDistanceKm", viewer_html)
            self.assertTrue(Path(rebuilt_viewer).exists())

    def test_project_paths_accept_fullwidth_screening_number_and_japanese_output(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp)
            route_dir = project / "10_ルート(Route)データ"
            trip_dir = project / "20_第２スクリーニング(ルート)"
            out_dir = project / "30_ルートパフォーマンス"
            route_dir.mkdir(parents=True)
            trip_dir.mkdir()
            out_dir.mkdir()
            write_route(route_dir / "route_a.csv")

            resolved_trip, resolved_route, resolved_out = route_performance.resolve_project_paths(project)

            self.assertEqual(resolved_trip, trip_dir)
            self.assertEqual(resolved_route, route_dir)
            self.assertEqual(resolved_out, out_dir)


if __name__ == "__main__":
    unittest.main()
