import argparse
import os
import re
import sys
import traceback
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
try:
    from openpyxl.drawing.image import Image as XLImage
except Exception as _exc:
    XLImage = None
    print(f"[WARN] openpyxl image feature disabled (Pillow missing?): {_exc}")
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins

print(
    "[BOOT] 32 start "
    f"py={sys.executable} "
    f"ver={sys.version.split()[0]} "
    f"cwd={os.getcwd()} "
    f"argv={' '.join(sys.argv)}",
    flush=True,
)

BATCH_JOBS: list[dict] = []
BATCH_MODE_ACTIVE = False
FOLDER_CROSS = "11_交差点(Point)データ"
FOLDER_31OUT = "31_交差点パフォーマンス"
FOLDER_32OUT = "32_交差点レポート"

COL_FILE = "抽出CSVファイル名"
COL_DATE = "運行日"
COL_VTYPE = "自動車の種別"
COL_USE = "用途"
COL_IN_BRANCH = "流入枝番"
COL_OUT_BRANCH = "流出枝番"
COL_DIST = "計測距離(m)"
COL_TIME = "所要時間(s)"
COL_T0 = "閑散時所要時間(s)"
COL_DELAY = "遅れ時間(s)"
COL_TIME_VALID = "所要時間算出可否"
COL_TIME_REASON = "所要時間算出不可理由"
COL_TIME_PRIMARY = "計測開始_GPS時刻(補間)"
COL_TIME_FALLBACK = "算出中心_GPS時刻"

DELAY_BINS = [
    (0, 5),
    (5, 10),
    (10, 20),
    (20, 30),
    (30, 60),
    (60, 120),
    (120, 180),
    (180, None),
]
DELAY_LABELS = ["0-5", "5-10", "10-20", "20-30", "30-60", "60-120", "120-180", "180+"]
TIME_LABELS = ["1-4時", "4-7時", "7-10時", "10-13時", "13-16時", "16-19時", "19-22時", "22-1時"]
MAP_SCALE = 0.26
MAP_ANCHOR_CELL = "B11"


def parse_center_datetime(val) -> datetime | None:
    if val is None:
        return None
    if pd.isna(val):
        return None
    text = str(val).strip()
    if not text:
        return None

    patterns = [
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y%m%d%H%M%S",
        "%Y%m%d%H%M%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%H:%M:%S",
        "%H:%M",
    ]

    for fmt in patterns:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def floor_to_30min(dt: datetime) -> datetime:
    return dt.replace(minute=(dt.minute // 30) * 30, second=0, microsecond=0)


def format_halfhour_label(slot_start: datetime) -> str:
    slot_end = slot_start + pd.Timedelta(minutes=30)
    return f"{slot_start.strftime('%H:%M')}～{slot_end.strftime('%H:%M')}"


def shorten_halfhour_label(label: str | None) -> str:
    if not label:
        return "-"
    start = str(label).split("～", 1)[0].strip()
    if not start:
        return "-"
    if re.fullmatch(r"\d{2}:\d{2}", start) and start.startswith("0"):
        start = start[1:]
    return f"{start}～"


def build_fixed_halfhour_slots() -> list[dict]:
    slots: list[dict] = []
    for slot_idx in range(48):
        start_total_min = slot_idx * 30
        end_total_min = start_total_min + 30
        start_h, start_m = divmod(start_total_min, 60)
        end_h, end_m = divmod(end_total_min, 60)
        label = f"{start_h}:{start_m:02d}～{end_h}:{end_m:02d}"
        slots.append({"slot_idx": slot_idx, "label": label})
    return slots


def hour_to_time_bin(hour: int) -> int:
    if hour in (22, 23, 0):
        return 7
    idx = (hour - 1) // 3
    return max(0, min(idx, 6))


def parse_operation_date(value: str):
    text = str(value).strip()
    if not text:
        return None
    digits = re.sub(r"\D", "", text)
    if len(digits) != 8:
        return None
    try:
        return datetime.strptime(digits, "%Y%m%d").date()
    except ValueError:
        return None


class _ExcelReportHelper:
    def __init__(
        self,
        crossroad_path: Path,
        image_path: Path,
        performance_path: Path,
        all_df: pd.DataFrame,
        clean_df: pd.DataFrame,
        unique_dates: list[date],
    ) -> None:
        self.crossroad_path = crossroad_path
        self.image_path = image_path
        self.performance_path = performance_path
        self.all_df = all_df
        self.clean_df = clean_df
        self.unique_dates = unique_dates

    def create(self, output_xlsx: Path) -> None:
        self._create_excel_report(output_xlsx)

    def _create_excel_report(self, save_path: Path) -> None:
        wb = Workbook()
        ws_report = wb.active
        ws_report.title = "Report"
        ws_delay = wb.create_sheet("遅れ時間（データ）")
        ws_time = wb.create_sheet("時間帯（データ）")

        self._configure_report_sheet(ws_report)
        combos = self._collect_combination_data()
        self._populate_delay_data_sheet(ws_delay, combos)
        self._populate_time_data_sheet(ws_time, combos)
        self._populate_report_sheet(ws_report, combos)

        wb.save(save_path)

    def _configure_report_sheet(self, ws) -> None:
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(
            left=0.7874,
            right=0.5906,
            top=0.7480,
            bottom=0.5512,
            header=ws.page_margins.header,
            footer=ws.page_margins.footer,
        )
        ws.print_options.horizontalCentered = True
        ws.print_title_rows = "1:1"
        ws.column_dimensions["A"].width = 16.0
        ws.column_dimensions["B"].width = 11.86
        ws.column_dimensions["C"].width = 9.5
        for col in ["D", "F", "G", "H", "I", "J", "K"]:
            ws.column_dimensions[col].width = 7.0
        ws.column_dimensions["E"].width = 8.5

    def _collect_combination_data(self) -> list[dict]:
        total_days = len(self.unique_dates)
        combos: list[dict] = []
        grouped = self.clean_df.groupby(["in_b", "out_b"])
        for (in_b, out_b), subset in grouped:
            count_total = len(subset)
            daily_count = count_total / total_days if total_days else 0
            ok_subset = subset[subset["time_valid"] == 1]
            avg_delay = ok_subset["delay_s"].mean() if not ok_subset.empty else 0
            total_delay = ok_subset["delay_s"].sum() if not ok_subset.empty else 0
            daily_total_delay_s = total_delay / total_days if total_days else 0
            daily_total_delay_min = daily_total_delay_s / 60 if total_days else 0
            time_per_day, time_parse_ng_count, time_bin_total = self._calc_time_per_day_counts(
                subset["time"], total_days
            )
            halfhour_summary = self._build_halfhour_summary(ok_subset, total_days)
            total_halfhour_delay_s = sum(item["delay_total_s"] for item in halfhour_summary)
            peak_slot = max(
                halfhour_summary,
                key=lambda item: (item["delay_total_s"], -item["slot_start"].timestamp()),
                default=None,
            )
            ok_count = len(ok_subset)
            ok_per_day = ok_count / total_days if total_days else 0
            time_bins_total_per_day = sum(time_per_day)
            print(
                "[CHECK] direction="
                f"{int(in_b)}→{int(out_b)} "
                f"daily_count={daily_count:.6f} "
                f"sum_time_bins_per_day={time_bins_total_per_day:.6f} "
                f"ok_per_day={ok_per_day:.6f} "
                f"daily_total_delay_s={daily_total_delay_s:.6f} "
                f"daily_total_delay_min={daily_total_delay_min:.6f} "
                f"avg_delay={avg_delay:.6f} "
                f"time_parse_ng_count={time_parse_ng_count} "
                f"time_bin_total={time_bin_total}"
            )

            combos.append(
                {
                    "in_b": int(in_b),
                    "out_b": int(out_b),
                    "count_total": count_total,
                    "ok_count": ok_count,
                    "daily_count": daily_count,
                    "avg_delay": avg_delay,
                    "total_delay": total_delay,
                    "daily_total_delay": daily_total_delay_s,
                    "time_per_day": time_per_day,
                    "halfhour_summary": halfhour_summary,
                    "total_halfhour_delay_s": total_halfhour_delay_s,
                    "peak_slot_label": peak_slot["slot_label"] if peak_slot else None,
                    "peak_slot_delay_s": peak_slot["delay_total_s"] if peak_slot else 0.0,
                    "peak_slot_delay_min": (peak_slot["delay_total_s"] / 60.0) if peak_slot else 0.0,
                    "peak_slot_avg_delay_s": peak_slot["delay_avg_s"] if peak_slot else 0.0,
                    "peak_slot_share_pct": (
                        (peak_slot["delay_total_s"] / total_halfhour_delay_s * 100.0)
                        if peak_slot and total_halfhour_delay_s
                        else 0.0
                    ),
                }
            )

        combos.sort(key=lambda x: (-x["daily_total_delay"], x["in_b"], x["out_b"]))
        return combos

    def _build_halfhour_summary(self, ok_subset: pd.DataFrame, total_days: int) -> list[dict]:
        slot_map: dict[datetime, dict] = {}
        for row in ok_subset[["time", "delay_s"]].itertuples(index=False):
            delay_s = pd.to_numeric(row.delay_s, errors="coerce")
            if pd.isna(delay_s):
                continue
            dt = parse_center_datetime(row.time)
            if dt is None:
                continue
            slot_start = floor_to_30min(dt)
            slot = slot_map.setdefault(slot_start, {"delay_total_s": 0.0, "count": 0})
            slot["delay_total_s"] += float(delay_s)
            slot["count"] += 1

        summary = []
        for slot_start in sorted(slot_map.keys()):
            delay_total_s = slot_map[slot_start]["delay_total_s"]
            count = slot_map[slot_start]["count"]
            summary.append(
                {
                    "slot_start": slot_start,
                    "slot_label": format_halfhour_label(slot_start),
                    "delay_total_s": delay_total_s,
                    "delay_total_min": delay_total_s / 60.0,
                    "delay_avg_s": (delay_total_s / count) if count else 0.0,
                    "count": count,
                    "daily_delay_total_s": (delay_total_s / total_days) if total_days else 0.0,
                    "daily_delay_total_min": (delay_total_s / total_days / 60.0) if total_days else 0.0,
                }
            )
        return summary

    def _calc_delay_per_day_counts(self, delay_series: pd.Series, total_days: int) -> list[float]:
        delays = pd.to_numeric(delay_series, errors="coerce").dropna().astype(float).tolist()
        counts = [0 for _ in DELAY_BINS]
        for v in delays:
            for idx, (low, high) in enumerate(DELAY_BINS):
                if high is None:
                    if v >= low:
                        counts[idx] += 1
                        break
                elif low <= v < high:
                    counts[idx] += 1
                    break
        if total_days == 0:
            return [0.0 for _ in DELAY_BINS]
        return [c / total_days for c in counts]

    def _calc_time_per_day_counts(self, time_series: pd.Series, total_days: int) -> tuple[list[float], int, int]:
        counts = [0 for _ in TIME_LABELS]
        time_parse_ng_count = 0
        for value in time_series.tolist():
            dt = parse_center_datetime(value)
            if dt is None:
                time_parse_ng_count += 1
                continue
            hour = dt.hour
            bin_idx = hour_to_time_bin(hour)
            counts[bin_idx] += 1
        if total_days == 0:
            return [0.0 for _ in TIME_LABELS], time_parse_ng_count, sum(counts)
        return [c / total_days for c in counts], time_parse_ng_count, sum(counts)

    def _populate_delay_data_sheet(self, ws, combos: list[dict]) -> None:
        headers = [
            "流入方向",
            "流出方向",
            "日あたり総遅れ時間（分/日）",
            "平均遅れ時間（秒）",
            "時間帯",
            "30分総遅れ時間（分）",
            "30分平均遅れ時間（秒）",
        ]
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)

        fixed_slots = build_fixed_halfhour_slots()
        total_days = len(self.unique_dates)
        delay_df = self.clean_df.copy()
        delay_df = delay_df[delay_df["time_valid"] == 1].copy()
        delay_df["delay_s_num"] = pd.to_numeric(delay_df["delay_s"], errors="coerce")
        delay_df = delay_df[delay_df["delay_s_num"].notna()].copy()
        delay_df = delay_df[delay_df["in_b"] != delay_df["out_b"]].copy()

        slot_indices: list[int | None] = []
        for value in delay_df["time"].tolist():
            dt = parse_center_datetime(value)
            if dt is None:
                slot_indices.append(None)
                continue
            slot_indices.append(dt.hour * 2 + (dt.minute // 30))
        delay_df["slot_idx"] = slot_indices
        delay_df = delay_df[delay_df["slot_idx"].notna()].copy()
        delay_df["slot_idx"] = delay_df["slot_idx"].astype(int)

        slot_stats_map: dict[tuple[int, int], dict[int, dict[str, float]]] = {}
        if not delay_df.empty:
            grouped = delay_df.groupby(["in_b", "out_b", "slot_idx"])
            for (in_b, out_b, slot_idx), subset in grouped:
                delay_total_s = float(subset["delay_s_num"].sum())
                count = int(len(subset))
                slot_stats_map.setdefault((int(in_b), int(out_b)), {})[int(slot_idx)] = {
                    "delay_total_min": delay_total_s / 60.0,
                    "delay_avg_s": (delay_total_s / count) if count else 0.0,
                }

        sorted_combos = sorted(
            [c for c in combos if c["in_b"] != c["out_b"]],
            key=lambda c: (c["in_b"], c["out_b"]),
        )
        for combo in sorted_combos:
            in_b = int(combo["in_b"])
            out_b = int(combo["out_b"])
            daily_total_delay_min = (combo["total_delay"] / total_days / 60.0) if total_days else 0.0
            avg_delay_s = float(combo["avg_delay"] or 0.0)
            direction_slots = slot_stats_map.get((in_b, out_b), {})
            for slot in fixed_slots:
                slot_data = direction_slots.get(slot["slot_idx"], None)
                ws.append(
                    [
                        in_b,
                        out_b,
                        daily_total_delay_min,
                        avg_delay_s,
                        slot["label"],
                        slot_data["delay_total_min"] if slot_data else 0.0,
                        slot_data["delay_avg_s"] if slot_data else 0.0,
                    ]
                )

        for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
            for cell in row:
                cell.number_format = "0.0"
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
            for cell in row:
                cell.number_format = "0.0"

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 17
        ws.column_dimensions["G"].width = 17

    def _populate_time_data_sheet(self, ws, combos: list[dict]) -> None:
        headers = [
            "方向（流入→流出）",
            "総台数（台）",
            "日あたり台数（台/日）",
            "平均遅れ時間（秒）",
            "1日あたり総遅れ時間（秒/日）",
            "階級（時）",
            "台数（台/日）",
        ]
        ws.append(headers)
        sorted_combos = sorted(combos, key=lambda c: (c["in_b"], c["out_b"]))
        for combo in sorted_combos:
            direction = f"{combo['in_b']}→{combo['out_b']}"
            base_info = [
                direction,
                combo["count_total"],
                combo["daily_count"],
                combo["avg_delay"],
                combo["daily_total_delay"],
            ]
            for label, per_day in zip(TIME_LABELS, combo["time_per_day"]):
                ws.append(base_info + [label, per_day])

        for row in ws.iter_rows(min_row=2, min_col=3, max_col=5):
            for cell in row:
                cell.number_format = "0.0"
        for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
            for cell in row:
                cell.number_format = "0.0"

    def _populate_report_sheet(self, ws, combos: list[dict]) -> None:
        cross_name = self.crossroad_path.stem
        title_text = f"ETC2.0 交差点パフォーマンス調査：{cross_name}"
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = Font(size=16, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        title_cell.alignment = Alignment(horizontal="center")

        summary_start_row = 3
        self._write_summary_block(ws, summary_start_row)
        image_obj = self._create_resized_image()
        if image_obj:
            ws.add_image(image_obj, MAP_ANCHOR_CELL)

        combos_for_report = [c for c in combos if int(c["in_b"]) != int(c["out_b"])]
        combos_for_report.sort(key=lambda x: (-x["daily_total_delay"], int(x["in_b"]), int(x["out_b"])))

        time_section_title_row = 26
        self._write_section_title(ws, row=time_section_title_row, text="方向別トリップ数集計表")

        time_title_row = time_section_title_row + 1
        time_header_row = time_title_row + 1
        time_data_row = time_title_row + 2
        time_last_row = self._write_time_table_pdf_style(
            ws, combos_for_report, time_title_row, time_header_row, time_data_row
        )

        delay_section_title_row = time_last_row + 2
        self._write_section_title(ws, row=delay_section_title_row, text="方向別遅れ時間集計表")

        delay_title_row = delay_section_title_row + 1
        delay_header_row = delay_title_row + 1
        delay_data_row = delay_title_row + 2
        delay_last_row = self._write_delay_table_pdf_style(
            ws, combos_for_report, delay_title_row, delay_header_row, delay_data_row
        )
        self._write_delay_table_note(ws, row=delay_last_row + 1)

    @staticmethod
    def _write_section_title(ws, row: int, text: str) -> None:
        title_cell = ws.cell(row=row, column=1, value=text)
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

    @staticmethod
    def _write_delay_table_note(ws, row: int) -> None:
        note_text = (
            "※遅れ時間とは、スムーズに走行した場合と比べた時間ロスを指し、"
            "本表は全トリップにおける遅れ時間を方向別・時間帯別に集計したもので、"
            "交差点の効率化の検討に活用できる。"
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        note_cell = ws.cell(row=row, column=1, value=note_text)
        note_cell.font = Font(size=9)
        note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 36

    def _write_summary_block(self, ws, start_row: int) -> int:
        start_date, end_date = (None, None)
        if self.unique_dates:
            start_date = self.unique_dates[0]
            end_date = self.unique_dates[-1]

        weekday_map = {0: "月", 1: "火", 2: "水", 3: "木", 4: "金", 5: "土", 6: "日"}
        weekday_order = [0, 1, 2, 3, 4, 5, 6]
        weekdays = ""
        if self.unique_dates:
            unique_weekdays = sorted({d.weekday() for d in self.unique_dates}, key=weekday_order.index)
            weekdays = "・".join(weekday_map[d] for d in unique_weekdays)

        total_records = len(self.all_df)
        ok_records = int(self.all_df["time_valid"].sum()) if total_records else 0
        ng_records = total_records - ok_records
        branch_ok = int((self.all_df["in_b"].notna() & self.all_df["out_b"].notna()).sum()) if total_records else 0
        branch_ng = total_records - branch_ok
        date_range = self._format_date_range(start_date, end_date)
        total_days_text = f"{len(self.unique_dates)}日（{date_range}）" if date_range else f"{len(self.unique_dates)}日"
        info_pairs = [
            ("交差点定義ファイル", self.crossroad_path.name, None),
            ("パフォーマンスCSV", self.performance_path.name, None),
            ("総日数", total_days_text, None),
            ("対象曜日", weekdays, None),
            ("総レコード数（通過）（台）", total_records, None),
            ("枝判定 OK/NG（台）", f"{branch_ok} / {branch_ng}", None),
            ("所要時間算出 OK/NG（台）", f"{ok_records} / {ng_records}", None),
        ]

        for offset, (label, value, extra) in enumerate(info_pairs):
            row_idx = start_row + offset
            label_cell = ws.cell(row=row_idx, column=1, value=f"{label}:")
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(wrap_text=False)
            value_cell = ws.cell(row=row_idx, column=4, value=value)
            value_cell.alignment = Alignment(wrap_text=False)

        return start_row + len(info_pairs)

    @staticmethod
    def _format_date_range(start_date: date | None, end_date: date | None) -> str:
        if not start_date or not end_date:
            return ""
        if start_date == end_date:
            return start_date.strftime("%Y年%m月%d日")
        return f"{start_date.strftime('%Y年%m月%d日')}～{end_date.strftime('%Y年%m月%d日')}"

    def _write_time_table_pdf_style(
        self, ws, combos: list[dict], title_row: int, header_row: int, data_row: int
    ) -> int:
        max_col = 11
        ws.cell(row=title_row, column=1, value="")
        ws.cell(row=title_row, column=2, value="")
        ws.cell(row=title_row, column=3, value="")
        ws.merge_cells(start_row=title_row, start_column=4, end_row=title_row, end_column=max_col)
        title_cell = ws.cell(row=title_row, column=4, value="時間帯ヒストグラム（台/日）")
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            "方向\n（流入→流出）",
            "日あたり\n台数\n（台/日）",
            "24h/\n7-19時\n（昼夜率）",
            "1-4\n時",
            "4-7\n時",
            "7-10\n時",
            "10-13\n時",
            "13-16\n時",
            "16-19\n時",
            "19-22\n時",
            "22-1\n時",
        ]
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[header_row].height = 50

        row_idx = data_row
        for combo in combos:
            direction = f"{combo['in_b']}→{combo['out_b']}"
            daytime_total = sum(combo["time_per_day"][2:6])
            day_night_ratio = (
                combo["daily_count"] / daytime_total if daytime_total else None
            )
            values = [
                direction,
                round(combo["daily_count"], 1),
                round(day_night_ratio, 2) if day_night_ratio is not None else None,
                *[round(v, 1) for v in combo["time_per_day"]],
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = Alignment(horizontal="center" if col == 1 else "right", vertical="center")
                if col == 3:
                    cell.number_format = "0.00"
                if col == 2 or col >= 4:
                    cell.number_format = "0.0"
            row_idx += 1

        total_row = row_idx
        total_daily = sum(combo["daily_count"] for combo in combos)
        total_time_bins = [
            sum(combo["time_per_day"][idx] for combo in combos) for idx in range(len(TIME_LABELS))
        ]
        total_daytime = sum(total_time_bins[2:6])
        total_ratio = total_daily / total_daytime if total_daytime else None
        total_values = [
            "合計",
            round(total_daily, 1),
            round(total_ratio, 2) if total_ratio is not None else None,
            *[round(v, 1) for v in total_time_bins],
        ]
        for col, val in enumerate(total_values, start=1):
            cell = ws.cell(row=total_row, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col == 1 else "right", vertical="center")
            if col == 3:
                cell.number_format = "0.00"
            if col == 2 or col >= 4:
                cell.number_format = "0.0"
        ws.row_dimensions[total_row].height = 18

        self.apply_table_borders(ws, title_row, 1, total_row, max_col)
        self._apply_row_bottom_border(ws, header_row, 1, max_col)
        self._apply_row_bottom_border(ws, total_row, 1, max_col)
        return total_row

    def _write_delay_table_pdf_style(
        self, ws, combos: list[dict], title_row: int, header_row: int, data_row: int
    ) -> int:
        max_col = 11

        overall_slot_totals: dict[datetime, float] = {}
        for combo in combos:
            for item in combo.get("halfhour_summary", []):
                overall_slot_totals[item["slot_start"]] = overall_slot_totals.get(item["slot_start"], 0.0) + item["delay_total_s"]

        def pick_peak_slot(start_hour: int, end_hour: int) -> datetime | None:
            candidates = [
                (slot_start, total)
                for slot_start, total in overall_slot_totals.items()
                if start_hour <= slot_start.hour < end_hour
            ]
            if not candidates:
                return None
            return max(candidates, key=lambda x: (x[1], -x[0].timestamp()))[0]

        am_peak_slot = pick_peak_slot(5, 12)
        pm_peak_slot = pick_peak_slot(12, 24)
        am_peak_label = format_halfhour_label(am_peak_slot) if am_peak_slot else "-"
        pm_peak_label = format_halfhour_label(pm_peak_slot) if pm_peak_slot else "-"
        am_peak_total = overall_slot_totals.get(am_peak_slot, 0.0) if am_peak_slot else 0.0
        pm_peak_total = overall_slot_totals.get(pm_peak_slot, 0.0) if pm_peak_slot else 0.0

        ws.merge_cells(start_row=title_row, start_column=1, end_row=data_row, end_column=1)
        ws.cell(row=title_row, column=1, value="方向（流入→流出）")
        ws.merge_cells(start_row=title_row, start_column=2, end_row=title_row, end_column=4)
        ws.cell(row=title_row, column=2, value="日あたり集計")
        ws.merge_cells(start_row=title_row, start_column=5, end_row=title_row, end_column=7)
        ws.cell(row=title_row, column=5, value="方向別ピーク")
        ws.merge_cells(start_row=title_row, start_column=8, end_row=title_row, end_column=9)
        ws.cell(row=title_row, column=8, value="午前ピーク")
        ws.merge_cells(start_row=title_row, start_column=10, end_row=title_row, end_column=11)
        ws.cell(row=title_row, column=10, value="午後ピーク")

        ws.merge_cells(start_row=header_row, start_column=2, end_row=data_row, end_column=2)
        ws.cell(row=header_row, column=2, value="日あたり総遅れ時間\n（分・台/日）")
        ws.merge_cells(start_row=header_row, start_column=3, end_row=data_row, end_column=3)
        ws.cell(row=header_row, column=3, value="総遅れ時間\n構成率（％）")
        ws.merge_cells(start_row=header_row, start_column=4, end_row=data_row, end_column=4)
        ws.cell(row=header_row, column=4, value="平均遅れ時間\n（秒）")
        ws.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=7)
        ws.cell(row=header_row, column=5, value="30分間毎集計")
        ws.merge_cells(start_row=header_row, start_column=8, end_row=header_row, end_column=9)
        ws.cell(row=header_row, column=8, value=am_peak_label)
        ws.merge_cells(start_row=header_row, start_column=10, end_row=header_row, end_column=11)
        ws.cell(row=header_row, column=10, value=pm_peak_label)

        detail_headers = [
            "時間",
            "総遅れ時間（分）",
            "平均遅れ時間（秒）",
            "総遅れ時間（分）",
            "構成率（％）",
            "総遅れ時間（分）",
            "構成率（％）",
        ]
        for idx, text in enumerate(detail_headers, start=5):
            ws.cell(row=data_row, column=idx, value=text)

        for row in range(title_row, data_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[title_row].height = 24
        ws.row_dimensions[header_row].height = 24
        ws.row_dimensions[data_row].height = 44

        row_idx = data_row + 1
        total_daily_delay = sum(c["daily_total_delay"] for c in combos)
        for combo in combos:
            slot_delay_map = {item["slot_start"]: item["delay_total_s"] for item in combo.get("halfhour_summary", [])}
            direction = f"{combo['in_b']}→{combo['out_b']}"
            daily_share_pct = (combo["daily_total_delay"] / total_daily_delay * 100.0) if total_daily_delay else 0.0
            am_direction_delay = slot_delay_map.get(am_peak_slot, 0.0) if am_peak_slot else 0.0
            pm_direction_delay = slot_delay_map.get(pm_peak_slot, 0.0) if pm_peak_slot else 0.0
            values = [
                direction,
                round(combo["daily_total_delay"] / 60.0, 1),
                round(daily_share_pct, 1),
                round(combo["avg_delay"], 1),
                shorten_halfhour_label(combo.get("peak_slot_label")),
                round(combo.get("peak_slot_delay_min", 0.0), 1),
                round(combo.get("peak_slot_avg_delay_s", 0.0), 1),
                round(am_direction_delay / 60.0, 1),
                round((am_direction_delay / am_peak_total * 100.0), 1) if am_peak_total else None,
                round(pm_direction_delay / 60.0, 1),
                round((pm_direction_delay / pm_peak_total * 100.0), 1) if pm_peak_total else None,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = Alignment(horizontal="center" if col in (1, 5) else "right", vertical="center")
                if col in (2, 3, 4, 6, 7, 8, 9, 10, 11):
                    cell.number_format = "0.0"
            row_idx += 1

        total_row = row_idx
        total_daily_delay_min = sum(c["daily_total_delay"] for c in combos) / 60.0
        total_ok = sum(c.get("ok_count", 0) for c in combos)
        total_delay_s = sum(c.get("total_delay", 0.0) for c in combos)
        total_avg_delay = (total_delay_s / total_ok) if total_ok else 0.0
        total_values = [
            "合計or平均",
            round(total_daily_delay_min, 1),
            100.0,
            round(total_avg_delay, 1),
            "-",
            "-",
            "-",
            "-",
            "-",
            "-",
            "-",
        ]
        for col, val in enumerate(total_values, start=1):
            cell = ws.cell(row=total_row, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col in (1, 5) else "right", vertical="center")
            if col in (2, 3, 4):
                cell.number_format = "0.0"
        ws.row_dimensions[total_row].height = 18

        self.apply_table_borders(ws, title_row, 1, total_row, max_col)
        self._apply_row_bottom_border(ws, header_row, 1, max_col)
        self._apply_row_bottom_border(ws, total_row, 1, max_col)
        return total_row

    @staticmethod
    def apply_table_borders(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
        thin = Side(style="thin")
        medium = Side(style="medium")
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                left = medium if col == min_col else thin
                right = medium if col == max_col else thin
                top = medium if row == min_row else thin
                bottom = medium if row == max_row else thin
                ws.cell(row=row, column=col).border = Border(
                    left=left, right=right, top=top, bottom=bottom
                )

    @staticmethod
    def _apply_row_bottom_border(ws, row: int, min_col: int, max_col: int) -> None:
        medium = Side(style="medium")
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            existing = cell.border
            cell.border = Border(
                left=existing.left,
                right=existing.right,
                top=existing.top,
                bottom=medium,
            )

    def _create_resized_image(self) -> "XLImage | None":
        if XLImage is None:
            return None
        if not self.image_path.exists():
            return None
        image = XLImage(str(self.image_path))
        try:
            original_width = image.width
            original_height = image.height
        except Exception:
            return image

        if original_width and original_height:
            image.width = max(1, int(original_width * MAP_SCALE))
            image.height = max(1, int(original_height * MAP_SCALE))
        elif image.width and image.height:
            image.width = max(1, int(image.width * MAP_SCALE))
            image.height = max(1, int(image.height * MAP_SCALE))
        return image

def create_excel_report_headless(
    crossroad_csv: Path,
    crossroad_img: Path,
    performance_csv: Path,
    output_xlsx: Path,
) -> None:
    df_perf = pd.read_csv(performance_csv, encoding="shift_jis")

    encodings = ["shift_jis", "cp932", "utf-8"]
    df_cross = None
    for enc in encodings:
        try:
            df_cross = pd.read_csv(crossroad_csv, encoding=enc)
            break
        except Exception:
            continue
    if df_cross is None:
        raise RuntimeError("交差点定義ファイルの読み込みに失敗しました。")

    required_cols = [
        COL_FILE,
        COL_DATE,
        COL_VTYPE,
        COL_USE,
        COL_IN_BRANCH,
        COL_OUT_BRANCH,
        COL_DIST,
        COL_TIME,
        COL_T0,
        COL_DELAY,
        COL_TIME_VALID,
        COL_TIME_REASON,
        COL_TIME_PRIMARY,
        COL_TIME_FALLBACK,
    ]
    missing = [c for c in required_cols if c not in df_perf.columns]
    if missing:
        raise RuntimeError(f"必要な列が見つかりません: {', '.join(missing)}")

    date_series = df_perf[COL_DATE].astype(str).apply(parse_operation_date)
    in_branch = pd.to_numeric(df_perf[COL_IN_BRANCH], errors="coerce")
    out_branch = pd.to_numeric(df_perf[COL_OUT_BRANCH], errors="coerce")
    time_val = pd.to_numeric(df_perf[COL_TIME], errors="coerce")
    t0_val = pd.to_numeric(df_perf[COL_T0], errors="coerce")
    delay_val = pd.to_numeric(df_perf[COL_DELAY], errors="coerce")
    time_valid = pd.to_numeric(df_perf[COL_TIME_VALID], errors="coerce")

    t_primary = df_perf[COL_TIME_FALLBACK].fillna("").astype(str).str.strip()
    t_fallback = df_perf[COL_TIME_PRIMARY].fillna("").astype(str).str.strip()
    time_series = t_primary.where(t_primary != "", t_fallback)

    data_all = pd.DataFrame(
        {
            "date": date_series,
            "in_b": in_branch,
            "out_b": out_branch,
            "time_s": time_val,
            "t0_s": t0_val,
            "delay_s": delay_val,
            "time_valid": time_valid,
            "time": time_series,
        }
    )

    data_all["time_valid"] = pd.to_numeric(data_all["time_valid"], errors="coerce").fillna(0).astype(int)

    data_clean = data_all.dropna(subset=["date", "in_b", "out_b"]).copy()
    data_clean["in_b"] = data_clean["in_b"].astype(int)
    data_clean["out_b"] = data_clean["out_b"].astype(int)
    unique_dates = sorted({d for d in data_all["date"] if pd.notna(d)})

    helper = _ExcelReportHelper(
        crossroad_path=crossroad_csv,
        image_path=crossroad_img,
        performance_path=performance_csv,
        all_df=data_all,
        clean_df=data_clean,
        unique_dates=unique_dates,
    )
    helper.crossroad_df = df_cross
    helper.create(output_xlsx)

def run_batch(jobs: list[dict]) -> int:
    global BATCH_MODE_ACTIVE
    BATCH_MODE_ACTIVE = True

    print("=== 32_crossroad_report (batch mode) ===")
    print(f"jobs: {len(jobs)}")

    ok = 0
    skipped = 0
    failed = 0

    for idx, job in enumerate(jobs, start=1):
        try:
            crossroad_csv = Path(job["crossroad_csv"])
            crossroad_img = Path(job["crossroad_img"])
            performance_csv = Path(job["performance_csv"])

            # 出力先（省略時は performance_csv の隣に *_report.xlsx）
            if "output_xlsx" in job and str(job["output_xlsx"]).strip():
                output_xlsx = Path(job["output_xlsx"])
            else:
                output_xlsx = performance_csv.with_name(f"{performance_csv.stem}_report.xlsx")

            print(f"\n[{idx}/{len(jobs)}]")
            print(f"  crossroad_csv : {crossroad_csv}")
            print(f"  crossroad_img : {crossroad_img}")
            print(f"  performance   : {performance_csv}")
            print(f"  output_xlsx   : {output_xlsx}")

            # 存在チェック（足りない場合は次へ）
            missing = [p for p in [crossroad_csv, crossroad_img, performance_csv] if not p.exists()]
            if missing:
                print("  [SKIP] missing files:")
                for m in missing:
                    print(f"    - {m}")
                skipped += 1
                continue

            output_xlsx.parent.mkdir(parents=True, exist_ok=True)

            create_excel_report_headless(crossroad_csv, crossroad_img, performance_csv, output_xlsx)
            print("  [OK] saved excel")
            ok += 1

        except Exception as exc:
            print(f"  [ERROR] batch job failed: {exc}")
            print(traceback.format_exc())
            failed += 1
            continue

    print("\n=== batch finished ===")
    print(f"summary: ok={ok}, skipped={skipped}, failed={failed}, jobs={len(jobs)}")
    BATCH_MODE_ACTIVE = False

    if len(jobs) == 0:
        print("[ERROR] no jobs to process")
        return 2
    if failed > 0:
        print("[ERROR] batch finished with failures")
        return 1
    if ok == 0:
        print("[ERROR] batch finished with no successful reports")
        return 1
    return 0



def _list_crossroad_names(cross_dir: Path) -> list[str]:
    if not cross_dir.exists():
        return []
    return [p.stem for p in sorted(cross_dir.glob("*.csv"))]


def build_jobs_from_project(project_dir: Path, targets: list[str] | None) -> list[dict]:
    cross_dir = project_dir / FOLDER_CROSS
    perf_dir = project_dir / FOLDER_31OUT
    out_dir = project_dir / FOLDER_32OUT
    out_dir.mkdir(parents=True, exist_ok=True)

    names = targets if targets else _list_crossroad_names(cross_dir)
    jobs: list[dict] = []
    for name in names:
        crossroad_csv = cross_dir / f"{name}.csv"
        crossroad_img = cross_dir / f"{name}.jpg"
        performance_csv = perf_dir / f"{name}_performance.csv"
        output_xlsx = out_dir / f"{name}_report.xlsx"

        missing = [p for p in [crossroad_csv, crossroad_img, performance_csv] if not p.exists()]
        if missing:
            print(f"[SKIP] {name}: missing files")
            for m in missing:
                print(f"  - {m}")
            continue

        jobs.append(
            {
                "crossroad_csv": str(crossroad_csv),
                "crossroad_img": str(crossroad_img),
                "performance_csv": str(performance_csv),
                "output_xlsx": str(output_xlsx),
            }
        )
    return jobs



def main() -> None:
    parser = argparse.ArgumentParser(description="32_crossroad_report (headless only)")
    parser.add_argument("--project", type=str, help="プロジェクトフォルダ")
    parser.add_argument("--targets", nargs="*", help="交差点名（stem）")
    args = parser.parse_args()

    if args.project:
        jobs = build_jobs_from_project(Path(args.project).resolve(), args.targets)
        code = run_batch(jobs)
        sys.exit(code)

    if BATCH_JOBS:
        code = run_batch(BATCH_JOBS)
        sys.exit(code)

    print("[ERROR] headless only: specify --project or define BATCH_JOBS")
    sys.exit(2)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print("[ERROR] unhandled exception in 32_crossroad_report.py")
        print(f"[ERROR] {exc}")
        traceback.print_exc()
        raise
