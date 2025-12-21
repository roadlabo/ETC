import sys
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
import matplotlib.font_manager as font_manager
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor, QPixmap, QTextCharFormat
from PySide6.QtWidgets import (
    QListWidget,
    QListWidgetItem,
    QApplication,
    QCalendarWidget,
    QFileDialog,
    QGridLayout,
    QLabel,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QHBoxLayout,
)

preferred_fonts = ["Meiryo", "Yu Gothic", "MS Gothic"]
installed_fonts = {f.name for f in font_manager.fontManager.ttflist}
for font_name in preferred_fonts:
    if font_name in installed_fonts:
        plt.rcParams["font.family"] = font_name
        break
plt.rcParams["axes.unicode_minus"] = False

# Column indices for performance data
COL_FILE = 2
COL_DATE = 3
COL_VTYPE = 7
COL_USE = 8
COL_IN_BRANCH = 9
COL_OUT_BRANCH = 10
COL_DIST = 11
COL_TIME = 12
COL_SPEED = 13
COL_CENTER_TIME = 31  # AF列：中心点_GPS時刻

# Column indices for crossroad definition data
COL_BRANCH_NO = 3
COL_DIR_DEG = 5

SPEED_BINS = [
    (0, 10),
    (10, 20),
    (20, 30),
    (30, 40),
    (40, 50),
    (50, 60),
    (60, None),
]
SPEED_LABELS = ["0-10", "10-20", "20-30", "30-40", "40-50", "50-60", "60+"]
TIME_BINS = [(0, 3), (3, 6), (6, 9), (9, 12), (12, 15), (15, 18), (18, 21), (21, 24)]
TIME_LABELS = ["0-3", "3-6", "6-9", "9-12", "12-15", "15-18", "18-21", "21-24"]
COMBOS_PER_PAGE = 20


class ScaledPixmapLabel(QLabel):
    def __init__(self, pixmap: QPixmap | None = None, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._pixmap: QPixmap | None = pixmap
        self.setAlignment(Qt.AlignCenter)
        if pixmap:
            self.setScaledContents(False)

    def setPixmap(self, pixmap: QPixmap) -> None:  # type: ignore[override]
        self._pixmap = pixmap
        super().setPixmap(pixmap)
        self._update_scaled_pixmap()

    def resizeEvent(self, event) -> None:  # type: ignore[override]
        super().resizeEvent(event)
        self._update_scaled_pixmap()

    def _update_scaled_pixmap(self) -> None:
        if not self._pixmap:
            return
        scaled = self._pixmap.scaled(self.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        super().setPixmap(scaled)


class MatplotlibCanvas(FigureCanvas):
    def __init__(self, parent: QWidget | None = None) -> None:
        self.fig = Figure(figsize=(8, 4))
        super().__init__(self.fig)
        self.setParent(parent)

    def clear(self) -> None:
        self.fig.clear()


def parse_center_datetime(val) -> datetime | None:
    if val is None:
        return None
    if pd.isna(val):
        return None
    text = str(val).strip()
    if not text:
        return None

    patterns = [
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y%m%d%H%M%S",
        "%H:%M:%S",
        "%H:%M",
    ]

    for fmt in patterns:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


class CrossroadViewer(QMainWindow):
    def __init__(self, crossroad_csv: Path, crossroad_jpg: Path, performance_csv: Path) -> None:
        super().__init__()
        self.setWindowTitle("Crossroad Performance Viewer")
        self.crossroad_path = crossroad_csv
        self.image_path = crossroad_jpg
        self.performance_path = performance_csv

        self._last_table_row = -1

        self.performance_df = pd.DataFrame()
        self.crossroad_df = pd.DataFrame()
        self.clean_df = pd.DataFrame()
        self.grouped_df = pd.DataFrame()
        self.unique_dates: list[datetime.date] = []
        self.unique_qdates: list[QDate] = []

        self._setup_ui()
        self._load_and_prepare()

    def _setup_ui(self) -> None:
        main_widget = QWidget()
        self.setCentralWidget(main_widget)

        main_layout = QVBoxLayout(main_widget)
        header_layout = QVBoxLayout()
        top_bar = QHBoxLayout()
        self.export_button = QPushButton("エクセル出力")
        self.export_button.clicked.connect(self.export_to_excel)
        top_bar.addStretch(1)
        top_bar.addWidget(self.export_button)
        header_layout.addLayout(top_bar)

        self.crossroad_label = QLabel("Crossroad file: -")
        self.performance_label = QLabel("Performance file: -")
        self.total_days_label = QLabel("総日数: -")
        self.total_records_label = QLabel("総レコード数: -")

        header_layout.addWidget(self.crossroad_label)
        header_layout.addWidget(self.performance_label)
        header_layout.addWidget(self.total_days_label)
        header_layout.addWidget(self.total_records_label)
        main_layout.addLayout(header_layout)

        splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(splitter, stretch=1)

        # Left splitter with image and calendar
        left_splitter = QSplitter(Qt.Vertical)
        splitter.addWidget(left_splitter)

        self.image_label = ScaledPixmapLabel()
        self.image_label.setMinimumHeight(200)
        left_splitter.addWidget(self.image_label)

        # Calendar + date list (right half)
        cal_container = QWidget()
        cal_layout = QGridLayout(cal_container)
        cal_layout.setContentsMargins(0, 0, 0, 0)

        self.calendar = QCalendarWidget()
        # remove week numbers (vertical header)
        self.calendar.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)

        self.date_list = QListWidget()
        self.date_list.setMinimumWidth(180)
        self.date_list.itemClicked.connect(self._on_date_clicked)

        cal_layout.addWidget(self.calendar, 0, 0)
        cal_layout.addWidget(self.date_list, 0, 1)
        left_splitter.addWidget(cal_container)

        # Right splitter with table+side on top and graph below
        right_splitter = QSplitter(Qt.Vertical)
        splitter.addWidget(right_splitter)

        right_top_splitter = QSplitter(Qt.Horizontal)
        right_splitter.addWidget(right_top_splitter)

        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)

        title_label = QLabel("交差点パフォーマンス表")
        table_layout.addWidget(title_label)

        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels([
            "流入枝番",
            "流出枝番",
            "総台数",
            "日あたり台数",
            "平均速度(km/h)",
        ])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.cellClicked.connect(self._on_row_clicked)
        self.table.itemSelectionChanged.connect(self._on_table_selection_changed)
        table_layout.addWidget(self.table)
        right_top_splitter.addWidget(table_container)

        side_container = QWidget()
        side_layout = QVBoxLayout(side_container)
        side_layout.setContentsMargins(0, 0, 0, 0)
        file_list_title = QLabel("該当ファイル一覧")
        side_layout.addWidget(file_list_title)
        self.file_list = QListWidget()
        self.file_list.setMinimumWidth(420)
        self.file_list.itemClicked.connect(self._on_file_clicked)
        self.file_list.currentItemChanged.connect(self._on_file_current_changed)
        side_layout.addWidget(self.file_list, stretch=3)
        detail_title = QLabel("選択ファイル詳細")
        side_layout.addWidget(detail_title)
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        side_layout.addWidget(self.detail_text, stretch=2)
        right_top_splitter.addWidget(side_container)

        graph_container = QWidget()
        graph_layout = QVBoxLayout(graph_container)
        graph_layout.setContentsMargins(0, 0, 0, 0)
        self.canvas = MatplotlibCanvas()
        graph_layout.addWidget(self.canvas)
        right_splitter.addWidget(graph_container)

        right_top_splitter.setStretchFactor(0, 4)
        right_top_splitter.setStretchFactor(1, 2)
        right_splitter.setStretchFactor(0, 3)
        right_splitter.setStretchFactor(1, 2)

        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 1)

    def _load_and_prepare(self) -> None:
        try:
            self.performance_df = pd.read_csv(self.performance_path, encoding="shift_jis")
        except Exception as exc:
            self._show_error(f"交差点パフォーマンスデータの読み込みに失敗しました: {exc}")
            return

        self.crossroad_df = self._load_crossroad_definition()
        if self.crossroad_df is None:
            return

        success = self._prepare_data()
        if not success:
            return

        self._load_image()
        self._populate_header()
        self._populate_table()
        self._highlight_calendar()
        self._populate_date_list_and_jump()

    def _load_crossroad_definition(self) -> pd.DataFrame | None:
        encodings = ["shift_jis", "cp932", "utf-8"]
        for enc in encodings:
            try:
                return pd.read_csv(self.crossroad_path, encoding=enc)
            except Exception:
                continue
        self._show_error("交差点定義ファイルの読み込みに失敗しました。")
        return None

    def _prepare_data(self) -> bool:
        try:
            date_series = self.performance_df.iloc[:, COL_DATE].astype(str).apply(self._parse_date)
            in_branch = pd.to_numeric(self.performance_df.iloc[:, COL_IN_BRANCH], errors="coerce")
            out_branch = pd.to_numeric(self.performance_df.iloc[:, COL_OUT_BRANCH], errors="coerce")
            speed = pd.to_numeric(self.performance_df.iloc[:, COL_SPEED], errors="coerce")
            center_time = self.performance_df.iloc[:, COL_CENTER_TIME]

            data = pd.DataFrame({
                "date": date_series,
                "in_b": in_branch,
                "out_b": out_branch,
                "spd": speed,
            })
            data = data.dropna()
            data["center_time"] = center_time.loc[data.index].values

            data["in_b"] = data["in_b"].astype(int)
            data["out_b"] = data["out_b"].astype(int)

            self.clean_df = data
            self.unique_dates = sorted({d for d in data["date"]})
            # Cache QDate list for calendar/list usage
            self.unique_qdates = [
                QDate(d.year, d.month, d.day) for d in self.unique_dates
            ]

            total_days = len(self.unique_dates)
            grouped = data.groupby(["in_b", "out_b"]).agg(
                総台数=("spd", "size"),
                平均速度=("spd", "mean"),
            )
            if total_days > 0:
                grouped["日あたり台数"] = grouped["総台数"] / total_days
            else:
                grouped["日あたり台数"] = 0

            grouped = grouped.reset_index()
            grouped = grouped.sort_values(
                by=["総台数", "in_b", "out_b"],
                ascending=[False, True, True],
            )
            self.grouped_df = grouped
            return True
        except Exception as exc:
            self._show_error(f"データ処理に失敗しました: {exc}")
            return False

    def _populate_header(self) -> None:
        self.crossroad_label.setText(f"Crossroad file: {self.crossroad_path.name}")
        self.performance_label.setText(f"Performance file: {self.performance_path.name}")
        self.total_days_label.setText(f"総日数: {len(self.unique_dates)}")
        self.total_records_label.setText(f"総レコード数: {len(self.clean_df)}")

    def _populate_table(self) -> None:
        df = self.grouped_df
        self.table.setRowCount(len(df))
        for row, (_, rec) in enumerate(df.iterrows()):
            # Ensure integer display (avoid "3.0" -> int("3.0") crash)
            in_b = int(rec["in_b"])
            out_b = int(rec["out_b"])
            in_item = QTableWidgetItem(str(in_b))
            out_item = QTableWidgetItem(str(out_b))
            total_item = QTableWidgetItem(str(int(rec["総台数"])))
            daily_item = QTableWidgetItem(f"{rec['日あたり台数']:.2f}")
            avg_item = QTableWidgetItem(f"{rec['平均速度']:.2f}")

            for item in (in_item, out_item, total_item, daily_item, avg_item):
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

            self.table.setItem(row, 0, in_item)
            self.table.setItem(row, 1, out_item)
            self.table.setItem(row, 2, total_item)
            self.table.setItem(row, 3, daily_item)
            self.table.setItem(row, 4, avg_item)
        self.table.resizeColumnsToContents()

    def _highlight_calendar(self) -> None:
        highlight_format = QTextCharFormat()
        highlight_format.setBackground(QColor("pink"))
        for qd in self.unique_qdates:
            if qd.isValid():
                self.calendar.setDateTextFormat(qd, highlight_format)

    def _populate_date_list_and_jump(self) -> None:
        # Right half list: all existing dates
        self.date_list.clear()
        for d in self.unique_dates:
            self.date_list.addItem(d.strftime("%Y-%m-%d"))

        # Default month: jump to the first existing date's month
        if self.unique_qdates:
            first = self.unique_qdates[0]
            self.calendar.setSelectedDate(first)
            self.calendar.setCurrentPage(first.year(), first.month())

    def _on_date_clicked(self, item) -> None:
        text = item.text()
        qdate = QDate.fromString(text, "yyyy-MM-dd")
        if not qdate.isValid():
            return
        self.calendar.setSelectedDate(qdate)
        self.calendar.setCurrentPage(qdate.year(), qdate.month())

    def _on_row_clicked(self, row: int, column: int) -> None:  # noqa: ARG002
        self._update_for_pair_by_row(row)

    def _on_table_selection_changed(self) -> None:
        row = self.table.currentRow()
        if row == -1 or row == self._last_table_row:
            return
        self._update_for_pair_by_row(row)

    def _update_for_pair_by_row(self, row: int) -> None:
        try:
            in_b_item = self.table.item(row, 0)
            out_b_item = self.table.item(row, 1)
            if not in_b_item or not out_b_item:
                return
            # Extra-safe parse (in case text becomes "3.0" again in future)
            in_b = int(float(in_b_item.text()))
            out_b = int(float(out_b_item.text()))
            self._last_table_row = row
            self._draw_histogram(in_b, out_b)
            self._update_file_list(in_b, out_b)
        except Exception as exc:
            self._show_error(f"ヒストグラム描画に失敗しました: {exc}")

    def _draw_histogram(self, in_b: int, out_b: int) -> None:
        subset = self.clean_df[(self.clean_df["in_b"] == in_b) & (self.clean_df["out_b"] == out_b)]
        self.canvas.clear()
        fig = self.canvas.fig
        ax_speed = fig.add_subplot(1, 2, 1)
        ax_time = fig.add_subplot(1, 2, 2)

        if subset.empty:
            for ax in (ax_speed, ax_time):
                ax.axis("off")
                ax.text(0.5, 0.5, "データなし", ha="center", va="center")
            self.canvas.draw()
            return

        speeds = subset["spd"].dropna().astype(float).tolist()
        avg_speed = subset["spd"].mean()
        count = len(speeds)

        # Fixed bins as percentages:
        # 0-10,10-20,20-30,30-40,40-50,50-60,60+
        labels = ["0-10", "10-20", "20-30", "30-40", "40-50", "50-60", "60+"]
        counts = [0] * 7
        for v in speeds:
            if v < 10:
                counts[0] += 1
            elif v < 20:
                counts[1] += 1
            elif v < 30:
                counts[2] += 1
            elif v < 40:
                counts[3] += 1
            elif v < 50:
                counts[4] += 1
            elif v < 60:
                counts[5] += 1
            else:
                counts[6] += 1
        perc = [c * 100.0 / count for c in counts] if count else [0.0] * 7

        if speeds:
            ax_speed.bar(labels, perc)
            ax_speed.set_ylim(0, max(perc) * 1.2 if max(perc) > 0 else 1)
            ax_speed.set_ylabel("割合(%)")
            for i, p in enumerate(perc):
                ax_speed.text(i, p, f"{p:.1f}%", ha="center", va="bottom", fontsize=9)
        else:
            ax_speed.axis("off")
            ax_speed.text(0.5, 0.5, "速度データなし", ha="center", va="center")

        time_labels = ["0-3", "3-6", "6-9", "9-12", "12-15", "15-18", "18-21", "21-24"]
        time_counts = [0] * 8
        parsed_times = [
            dt for dt in subset["center_time"].apply(parse_center_datetime).tolist() if dt is not None
        ]
        for dt in parsed_times:
            hour = dt.hour
            if hour < 3:
                time_counts[0] += 1
            elif hour < 6:
                time_counts[1] += 1
            elif hour < 9:
                time_counts[2] += 1
            elif hour < 12:
                time_counts[3] += 1
            elif hour < 15:
                time_counts[4] += 1
            elif hour < 18:
                time_counts[5] += 1
            elif hour < 21:
                time_counts[6] += 1
            elif hour < 24:
                time_counts[7] += 1

        if parsed_times:
            time_total = len(parsed_times)
            time_perc = [c * 100.0 / time_total for c in time_counts]
            ax_time.bar(time_labels, time_perc)
            ax_time.set_ylim(0, max(time_perc) * 1.2 if max(time_perc) > 0 else 1)
            ax_time.set_ylabel("割合(%)")
            for i, p in enumerate(time_perc):
                ax_time.text(i, p, f"{p:.1f}%", ha="center", va="bottom", fontsize=9)
        else:
            ax_time.axis("off")
            ax_time.text(0.5, 0.5, "時刻データなし", ha="center", va="center")

        fig.suptitle(f"{in_b}→{out_b} / 台数:{count} / 平均速度:{avg_speed:.1f} km/h")
        fig.tight_layout(rect=[0, 0, 1, 0.92])
        self.canvas.draw()

    def _update_file_list(self, in_b: int, out_b: int) -> None:
        self.file_list.clear()
        self.detail_text.setPlainText("")
        try:
            in_series = pd.to_numeric(self.performance_df.iloc[:, COL_IN_BRANCH], errors="coerce")
            out_series = pd.to_numeric(self.performance_df.iloc[:, COL_OUT_BRANCH], errors="coerce")
            mask = (in_series == in_b) & (out_series == out_b)
            filtered = self.performance_df[mask]
            if filtered.empty:
                return

            file_series = filtered.iloc[:, COL_FILE].fillna("").astype(str)
            seen: set[str] = set()
            for idx, file_name in zip(filtered.index, file_series):
                if not file_name or file_name in seen:
                    continue
                seen.add(file_name)
                item = QListWidgetItem(file_name)
                item.setData(Qt.UserRole, idx)
                self.file_list.addItem(item)
        except Exception as exc:
            self._show_error(f"ファイル一覧の更新に失敗しました: {exc}")

    def _on_file_clicked(self, item: QListWidgetItem) -> None:
        self._update_file_detail(item)

    def _on_file_current_changed(
        self, current: QListWidgetItem | None, previous: QListWidgetItem | None
    ) -> None:  # noqa: ARG002
        if current is None:
            self.detail_text.setPlainText("")
            return
        self._update_file_detail(current)

    def _update_file_detail(self, item: QListWidgetItem) -> None:
        try:
            row_index = item.data(Qt.UserRole)
            if row_index is None:
                return
            row = self.performance_df.loc[row_index]

            vtype_map = {0: "軽二輪", 1: "大型", 2: "普通", 3: "小型", 4: "軽自動車"}
            use_map = {0: "未使用", 1: "乗用", 2: "貨物", 3: "特殊", 4: "乗合"}

            def _format_value(value, fmt: str = "{}") -> str:
                if pd.isna(value):
                    return "不明"
                return fmt.format(value)

            file_name = _format_value(row.iloc[COL_FILE])

            vtype_val = pd.to_numeric(row.iloc[COL_VTYPE], errors="coerce")
            vtype_int = int(vtype_val) if not pd.isna(vtype_val) else None
            vtype_text = vtype_map.get(vtype_int, "不明") if vtype_int is not None else "不明"
            vtype_display = f"{vtype_int}" if vtype_int is not None else "不明"

            use_val = pd.to_numeric(row.iloc[COL_USE], errors="coerce")
            use_int = int(use_val) if not pd.isna(use_val) else None
            use_text = use_map.get(use_int, "不明") if use_int is not None else "不明"
            use_display = f"{use_int}" if use_int is not None else "不明"

            dist_val = pd.to_numeric(row.iloc[COL_DIST], errors="coerce")
            time_val = pd.to_numeric(row.iloc[COL_TIME], errors="coerce")
            speed_val = pd.to_numeric(row.iloc[COL_SPEED], errors="coerce")

            detail_lines = [
                f"ファイル名：{file_name}",
                f"自動車の種別：{vtype_display}（{vtype_text}）",
                f"用途：{use_display}（{use_text}）",
                f"道なり距離(m)：{_format_value(dist_val, '{:.0f}')}",
                f"所要時間(s)：{_format_value(time_val, '{:.0f}')}",
                f"交差点通過速度(km/h)：{_format_value(speed_val, '{:.1f}')}",
            ]

            self.detail_text.setPlainText("\n".join(detail_lines))
        except Exception as exc:
            self._show_error(f"詳細表示の更新に失敗しました: {exc}")

    def _parse_date(self, value: str):
        text = str(value).strip()
        if not text:
            return None
        digits = re.sub(r"\D", "", text)
        if len(digits) != 8:
            return None
        try:
            return datetime.strptime(digits, "%Y%m%d").date()
        except ValueError:
            return None

    def _load_image(self) -> None:
        if not self.image_path.exists():
            return
        pixmap = QPixmap(str(self.image_path))
        if pixmap.isNull():
            return
        self.image_label.setPixmap(pixmap)

    def _show_error(self, message: str) -> None:
        QMessageBox.critical(self, "Error", message)

    def export_to_excel(self) -> None:
        if self.clean_df.empty:
            self._show_error("出力するデータがありません。先にファイルを読み込んでください。")
            return

        default_name = f"{self.crossroad_path.stem}_report.xlsx"
        save_path_str, _ = QFileDialog.getSaveFileName(
            self,
            "Excelレポートを保存",
            str(self.crossroad_path.with_name(default_name)),
            "Excel Files (*.xlsx)",
        )
        if not save_path_str:
            return

        try:
            self._create_excel_report(Path(save_path_str))
            QMessageBox.information(self, "完了", "エクセルレポートを出力しました。")
        except Exception as exc:  # pragma: no cover - UI path
            self._show_error(f"エクセル出力に失敗しました: {exc}")

    def _create_excel_report(self, save_path: Path) -> None:
        wb = Workbook()
        ws_report = wb.active
        ws_report.title = "Report"
        ws_data = wb.create_sheet("Data")

        self._configure_report_sheet(ws_report)
        combos = self._collect_combination_data()
        self._populate_data_sheet(ws_data, combos)
        self._populate_report_sheet(ws_report, combos)

        wb.save(save_path)

    def _configure_report_sheet(self, ws) -> None:
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.3
        ws.page_margins.bottom = 0.3
        ws.print_options.horizontalCentered = True
        widths = [6, 6, 8, 10, 10] + [6] * 20
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _collect_combination_data(self) -> list[dict]:
        total_days = len(self.unique_dates)
        combos: list[dict] = []
        grouped = self.clean_df.groupby(["in_b", "out_b"])
        for (in_b, out_b), subset in grouped:
            count_total = len(subset)
            daily_count = count_total / total_days if total_days else 0
            avg_speed = subset["spd"].mean() if not subset.empty else 0

            speed_perc = self._calc_speed_percent(subset["spd"])
            time_perc = self._calc_time_percent(subset["center_time"])

            combos.append(
                {
                    "in_b": int(in_b),
                    "out_b": int(out_b),
                    "count_total": count_total,
                    "daily_count": daily_count,
                    "avg_speed": avg_speed,
                    "speed_percent": speed_perc,
                    "time_percent": time_perc,
                }
            )

        combos.sort(key=lambda x: (-x["count_total"], x["in_b"], x["out_b"]))
        return combos

    def _calc_speed_percent(self, speed_series: pd.Series) -> list[float]:
        speeds = pd.to_numeric(speed_series, errors="coerce").dropna().astype(float).tolist()
        counts = [0 for _ in SPEED_BINS]
        for v in speeds:
            for idx, (low, high) in enumerate(SPEED_BINS):
                if high is None:
                    if v >= low:
                        counts[idx] += 1
                        break
                elif low <= v < high:
                    counts[idx] += 1
                    break
        total = sum(counts)
        if total == 0:
            return [0.0 for _ in SPEED_BINS]
        return [c * 100.0 / total for c in counts]

    def _calc_time_percent(self, time_series: pd.Series) -> list[float]:
        parsed = [parse_center_datetime(v) for v in time_series.tolist()]
        valid_hours = [dt.hour for dt in parsed if dt is not None]
        counts = [0 for _ in TIME_BINS]
        for hour in valid_hours:
            for idx, (low, high) in enumerate(TIME_BINS):
                if low <= hour < high:
                    counts[idx] += 1
                    break
        total = sum(counts)
        if total == 0:
            return [0.0 for _ in TIME_BINS]
        return [c * 100.0 / total for c in counts]

    def _populate_data_sheet(self, ws, combos: list[dict]) -> None:
        headers = [
            "in_b",
            "out_b",
            "count_total",
            "daily_count",
            "avg_speed",
            "metric_type",
            "bin_label",
            "percent",
        ]
        ws.append(headers)
        for combo in combos:
            base_info = [
                combo["in_b"],
                combo["out_b"],
                combo["count_total"],
                combo["daily_count"],
                combo["avg_speed"],
            ]
            for label, perc in zip(SPEED_LABELS, combo["speed_percent"]):
                ws.append(base_info + ["speed", label, perc])
            for label, perc in zip(TIME_LABELS, combo["time_percent"]):
                ws.append(base_info + ["time", label, perc])

        for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
            for cell in row:
                cell.number_format = "0.0"

    def _populate_report_sheet(self, ws, combos: list[dict]) -> None:
        current_row = 1
        combo_index = 0
        page_index = 0

        while combo_index < len(combos):
            current_row = self._write_page_header(ws, current_row)
            if page_index == 0:
                current_row = self._write_overview_table(ws, current_row)
                current_row += 2

            batch = combos[combo_index : combo_index + COMBOS_PER_PAGE]
            current_row = self._write_speed_table(ws, batch, current_row)
            current_row += 2
            current_row = self._write_time_table(ws, batch, current_row)

            combo_index += len(batch)
            page_index += 1
            if combo_index < len(combos):
                ws.row_breaks.append(Break(id=current_row))
                current_row += 3

    def _write_page_header(self, ws, start_row: int) -> int:
        title_cell = ws.cell(row=start_row, column=1, value="Crossroad Performance Report")
        title_cell.font = Font(size=16, bold=True)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=18)
        title_cell.alignment = Alignment(horizontal="center")

        info_pairs = [
            ("Crossroad file", self.crossroad_path.name),
            ("Performance file", self.performance_path.name),
            ("総日数", len(self.unique_dates)),
            ("総レコード数", len(self.clean_df)),
        ]
        info_row = start_row + 1
        info_col = 14
        for label, value in info_pairs:
            ws.cell(row=info_row, column=info_col, value=f"{label}:").font = Font(bold=True)
            ws.cell(row=info_row, column=info_col + 1, value=value)
            info_row += 1

        map_row = start_row + 2
        image_obj = self._create_resized_image()
        image_rows = 0
        if image_obj:
            ws.add_image(image_obj, f"A{map_row}")
            image_rows = max(12, int(image_obj.height / 18))
        return map_row + image_rows + 1

    def _write_overview_table(self, ws, start_row: int) -> int:
        headers = ["流入枝番", "流出枝番", "総台数", "日あたり台数", "平均速度(km/h)"]
        header_row = start_row
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=Side(style="thin"))

        for offset, rec in enumerate(self.grouped_df.itertuples(index=False), start=1):
            row_idx = header_row + offset
            ws.cell(row=row_idx, column=1, value=int(rec.in_b)).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=2, value=int(rec.out_b)).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=3, value=int(rec.総台数)).alignment = Alignment(horizontal="right")
            daily_cell = ws.cell(row=row_idx, column=4, value=float(rec.日あたり台数))
            daily_cell.alignment = Alignment(horizontal="right")
            daily_cell.number_format = "0.0"
            avg_cell = ws.cell(row=row_idx, column=5, value=float(rec.平均速度))
            avg_cell.alignment = Alignment(horizontal="right")
            avg_cell.number_format = "0.0"
        return header_row + len(self.grouped_df) + 1

    def _write_speed_table(self, ws, combos: list[dict], start_row: int) -> int:
        headers = ["in_b", "out_b", "count", "/day", "avg_spd"] + SPEED_LABELS
        header_row = start_row
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=Side(style="thin"))

        row_idx = header_row + 1
        for combo in combos:
            values = [
                combo["in_b"],
                combo["out_b"],
                combo["count_total"],
                combo["daily_count"],
                combo["avg_speed"],
                *[round(v, 1) for v in combo["speed_percent"]],
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                align = Alignment(horizontal="center") if col <= 2 else Alignment(horizontal="right")
                cell.alignment = align
                if col in (4, 5) or col > 5:
                    cell.number_format = "0.0"
            row_idx += 1
        return row_idx

    def _write_time_table(self, ws, combos: list[dict], start_row: int) -> int:
        headers = ["in_b", "out_b", "count", "/day", "avg_spd"] + TIME_LABELS
        header_row = start_row
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=Side(style="thin"))

        row_idx = header_row + 1
        for combo in combos:
            values = [
                combo["in_b"],
                combo["out_b"],
                combo["count_total"],
                combo["daily_count"],
                combo["avg_speed"],
                *[round(v, 1) for v in combo["time_percent"]],
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                align = Alignment(horizontal="center") if col <= 2 else Alignment(horizontal="right")
                cell.alignment = align
                if col in (4, 5) or col > 5:
                    cell.number_format = "0.0"
            row_idx += 1
        return row_idx

    def _create_resized_image(self) -> XLImage | None:
        if not self.image_path.exists():
            return None
        image = XLImage(str(self.image_path))
        try:
            original_width = image.width
            original_height = image.height
        except Exception:
            return image
        max_width = 900
        if original_width and original_width > max_width:
            ratio = max_width / float(original_width)
            image.width = max_width
            image.height = int(original_height * ratio)
        return image


def pick_three_files() -> tuple[Path, Path, Path] | None:
    while True:
        csv_path, _ = QFileDialog.getOpenFileName(
            None, "交差点ファイル（*.csv）を選択", "", "CSV (*.csv)"
        )
        if not csv_path:
            return None

        img_path, _ = QFileDialog.getOpenFileName(
            None,
            "交差点画像（*.jpg）を選択",
            str(Path(csv_path).parent),
            "Images (*.jpg *.jpeg *.png *.bmp)",
        )
        if not img_path:
            return None

        perf_path, _ = QFileDialog.getOpenFileName(
            None,
            "交差点パフォーマンス（*_performance.csv）を選択",
            str(Path(csv_path).parent),
            "CSV (*.csv)",
        )
        if not perf_path:
            return None

        csv_p = Path(csv_path)
        img_p = Path(img_path)
        perf_p = Path(perf_path)

        base = csv_p.stem
        ok = (img_p.stem == base) and (perf_p.stem == f"{base}_performance")
        if ok:
            return csv_p, img_p, perf_p

        QMessageBox.warning(
            None,
            "ファイル名が一致しません",
            f"選択ルール：\n"
            f"- {base}.csv\n"
            f"- {base}.jpg\n"
            f"- {base}_performance.csv\n\n"
            f"もう一度選び直してください。",
        )


def main() -> None:
    app = QApplication(sys.argv)
    picked = pick_three_files()
    if picked is None:
        sys.exit(0)
    crossroad_csv, crossroad_jpg, performance_csv = picked
    viewer = CrossroadViewer(crossroad_csv, crossroad_jpg, performance_csv)
    viewer.resize(1200, 800)
    viewer.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
